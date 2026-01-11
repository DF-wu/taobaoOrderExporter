#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
订单匹配和上色工具
"""
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
from datetime import datetime
import sys
import io

# 设置输出编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 文件路径
OLD_FILE = "淘寶暫時清單 2025.11.15.xlsx"
NEW_FILE = "2026淘寶結賬.xlsx"

# 颜色定义
COLOR_MOM = "FFE6E6"      # 淡红色（媽媽页）
COLOR_OTHER = "FFFFE0"    # 淡黄色（其他页）
COLOR_DUPLICATE = "FF0000" # 大红色（重复出现）

def create_backup(file_path):
    """创建备份文件"""
    from shutil import copy2

    backup_name = f"{Path(file_path).stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{Path(file_path).suffix}"
    backup_path = Path(file_path).parent / backup_name

    try:
        copy2(file_path, backup_path)
        print(f"✓ 备份已创建: {backup_path}")
        return str(backup_path)
    except Exception as e:
        print(f"✗ 备份失败: {e}")
        return None

def read_order_numbers(file_path, sheet_name):
    """读取指定工作表的订单编号"""
    print(f"\n读取 '{sheet_name}' 页的订单编号...")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[sheet_name]

    order_numbers = set()

    # 找到订单编号列（假设在第2列，索引为1）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1]:  # 订单编号在第2列（索引1）
            order_num = str(row[1]).strip()
            if order_num:
                order_numbers.add(order_num)

    wb.close()
    print(f"  找到 {len(order_numbers)} 个订单编号")
    return order_numbers

def process_new_file(new_file, mom_orders, other_orders):
    """处理新表并上色"""
    print(f"\n处理新表: {new_file}")

    wb = openpyxl.load_workbook(new_file)
    sheet = wb['2025淘寶']

    # 统计
    count_mom = 0
    count_other = 0
    count_duplicate = 0

    # 从第2行开始（跳过标题）
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        order_num = row[1].value  # 订单号在第2列（索引1）

        if order_num:
            order_num_str = str(order_num).strip()

            # 检查是否在两个集合中都存在
            in_mom = order_num_str in mom_orders
            in_other = order_num_str in other_orders

            if in_mom and in_other:
                # 同时存在于两页 - 大红色
                fill = PatternFill(start_color=COLOR_DUPLICATE,
                                 end_color=COLOR_DUPLICATE,
                                 fill_type="solid")
                count_duplicate += 1
            elif in_mom:
                # 只在媽媽页 - 淡红色
                fill = PatternFill(start_color=COLOR_MOM,
                                 end_color=COLOR_MOM,
                                 fill_type="solid")
                count_mom += 1
            elif in_other:
                # 只在其他页 - 淡黄色
                fill = PatternFill(start_color=COLOR_OTHER,
                                 end_color=COLOR_OTHER,
                                 fill_type="solid")
                count_other += 1
            else:
                continue

            # 对整行上色
            for cell in row:
                cell.fill = fill

    # 保存文件
    output_file = f"{Path(new_file).stem}_已上色{Path(new_file).suffix}"
    wb.save(output_file)
    wb.close()

    print(f"\n✓ 处理完成!")
    print(f"  淡红色（媽媽页）: {count_mom} 个订单")
    print(f"  淡黄色（其他页）: {count_other} 个订单")
    print(f"  大红色（重复）: {count_duplicate} 个订单")
    print(f"\n✓ 已保存到: {output_file}")

    return output_file

def main():
    print("="*60)
    print("订单匹配和上色工具")
    print("="*60)

    # 1. 创建备份
    print("\n步骤 1: 创建备份")
    backup_path = create_backup(NEW_FILE)
    if not backup_path:
        print("备份失败，终止操作")
        return

    # 2. 读取旧表订单编号
    print("\n步骤 2: 读取旧表订单编号")
    mom_orders = read_order_numbers(OLD_FILE, '媽媽')
    other_orders = read_order_numbers(OLD_FILE, '其他')

    # 检查是否有重复
    duplicates = mom_orders & other_orders
    if duplicates:
        print(f"\n⚠ 警告: 发现 {len(duplicates)} 个订单同时存在于'媽媽'和'其他'页")
        print(f"  这些订单将用大红色标记")

    # 3. 处理新表
    print("\n步骤 3: 匹配订单并上色")
    output_file = process_new_file(NEW_FILE, mom_orders, other_orders)

    print("\n" + "="*60)
    print("处理完成！")
    print("="*60)

if __name__ == "__main__":
    main()

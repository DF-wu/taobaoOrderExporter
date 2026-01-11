#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
安全的XLSX文件检查工具
"""
import openpyxl
import pandas as pd
from pathlib import Path
import sys
import io

# 设置输出编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def inspect_xlsx(file_path):
    """安全地检查xlsx文件信息"""
    print(f"\n{'='*60}")
    print(f"文件: {file_path}")
    print(f"{'='*60}")

    try:
        # 使用openpyxl加载（不修改）
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        print(f"\n工作表数量: {len(wb.sheetnames)}")
        print(f"工作表列表: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n--- 工作表: {sheet_name} ---")
            print(f"  行数: {sheet.max_row}")
            print(f"  列数: {sheet.max_column}")

            # 显示前5行数据预览
            print(f"  前5行预览:")
            for i, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), 1):
                print(f"    行{i}: {row}")

        wb.close()
        return True

    except Exception as e:
        print(f"❌ 错误: {e}")
        return False

def create_backup(file_path):
    """创建备份文件"""
    from shutil import copy2
    from datetime import datetime

    backup_name = f"{Path(file_path).stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{Path(file_path).suffix}"
    backup_path = Path(file_path).parent / backup_name

    try:
        copy2(file_path, backup_path)
        print(f"✓ 备份已创建: {backup_path}")
        return str(backup_path)
    except Exception as e:
        print(f"❌ 备份失败: {e}")
        return None

if __name__ == "__main__":
    # 检查当前目录的所有xlsx文件
    current_dir = Path(".")
    xlsx_files = list(current_dir.glob("*.xlsx"))

    # 过滤掉临时文件
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

    print(f"\n找到 {len(xlsx_files)} 个xlsx文件")

    for xlsx_file in xlsx_files:
        inspect_xlsx(xlsx_file)

    print(f"\n{'='*60}")
    print("检查完成")
    print(f"{'='*60}\n")
